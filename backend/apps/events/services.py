import io
import logging
from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from apps.common.background import fire_and_forget
from .tasks import send_event_registration_confirmation

logger = logging.getLogger(__name__)


def register_participant(
    event, user, email: str, first_name: str, last_name: str,
    nationality: str, organisation: str, profession: str, motivation: str,
) -> "EventParticipant":
    from .models import EventParticipant

    with transaction.atomic():
        event_locked = type(event).objects.select_for_update().get(pk=event.pk)

        if event_locked.registration_deadline and timezone.now() > event_locked.registration_deadline:
            raise ValidationError("Les inscriptions sont closes pour cet événement.")

        if event_locked.is_full:
            raise ValidationError("Cet événement est complet.")

        participant, created = EventParticipant.objects.get_or_create(
            event=event_locked,
            email=email,
            defaults={
                "user": user,
                "first_name": first_name,
                "last_name": last_name,
                "nationality": nationality,
                "organisation": organisation,
                "profession": profession,
                "motivation": motivation,
            },
        )
        if not created:
            raise ValidationError("Vous êtes déjà inscrit à cet événement.")

        transaction.on_commit(
            lambda: fire_and_forget(
                send_event_registration_confirmation.delay, participant.pk,
                error_message=f"Impossible d'envoyer l'email de confirmation à {email}",
            )
        )

        logger.info("Inscription : %s → %s", email, event.title)
        return participant


def find_latest_participant_info(email: str) -> "EventParticipant | None":
    from .models import EventParticipant

    return (
        EventParticipant.objects.filter(email__iexact=email)
        .order_by("-created_at")
        .first()
    )


def validate_presence(participant) -> "EventParticipant":
    if participant.presence_validated:
        raise ValidationError("La présence a déjà été validée.")

    participant.presence_validated = True
    participant.attended_at = timezone.now()
    participant.save(update_fields=["presence_validated", "attended_at"])
    return participant


def generate_qr_code(event) -> None:
    import qrcode
    from django.core.files.base import ContentFile
    from django.conf import settings

    url = f"{settings.FRONTEND_URL}/events/{event.pk}"
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    event.qr_code.save(f"qr_{event.pk}.png", ContentFile(buffer.getvalue()), save=True)


def export_participants_excel(event) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = [
        "Prénom", "Nom", "Email", "Nationalité", "Organisation", "Profession",
        "Raison de l'inscription", "Inscrit le", "Présence validée", "Heure d'arrivée",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, p in enumerate(event.participants.all(), 2):
        ws.cell(row=row, column=1, value=p.first_name)
        ws.cell(row=row, column=2, value=p.last_name)
        ws.cell(row=row, column=3, value=p.email)
        ws.cell(row=row, column=4, value=p.nationality)
        ws.cell(row=row, column=5, value=p.organisation)
        ws.cell(row=row, column=6, value=p.profession)
        ws.cell(row=row, column=7, value=p.motivation)
        ws.cell(row=row, column=8, value=p.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=9, value="Oui" if p.presence_validated else "Non")
        ws.cell(row=row, column=10, value=p.attended_at.strftime("%d/%m/%Y %H:%M") if p.attended_at else "")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
